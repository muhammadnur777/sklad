from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.http import JsonResponse
from .models import Product, Category, Unit
from finance.models import Purchase, PurchaseItem, StockMovement
import json
from django.views.decorators.http import require_POST
from django.db.models import F
from finance.models import Sale, SaleItem, StockMovement
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from django.http import HttpResponse
from finance.models import BazarSaleItem, Shop
from django.db.models import Sum
from datetime import date
from finance.models import BazarSale, Shop, Purchase
from django.db.models import Sum
from datetime import date, timedelta
from finance.models import BazarSale, Shop
from django.db.models import Sum
from datetime import date


@login_required(login_url='login')
def product_list(request):
    products = Product.objects.filter(is_active=True).select_related('category', 'unit')
    categories = Category.objects.all()

    category_id = request.GET.get('category')
    low_stock = request.GET.get('low')

    if category_id:
        products = products.filter(category_id=category_id)

    if low_stock:
        products = [p for p in products if p.is_low_stock and p.min_stock > 0]
    search = request.GET.get('search', '')
    if search:
        products = products.filter(name__icontains=search)

    context = {
        'products': products,
        'categories': categories,
        'current_category': category_id,
        'search': search,
        'low_stock': low_stock,
    }
    return render(request, 'inventory/product_list.html', context)


@login_required(login_url='login')
def add_product(request):
    categories = Category.objects.all()
    units = Unit.objects.all()

    if request.method == 'POST':
        name = request.POST.get('name')
        category_id = request.POST.get('category')
        unit_id = request.POST.get('unit')
        sell_price = int(request.POST.get('sell_price', 0) or 0)
        stock_boxes = int(request.POST.get('stock', 0) or 0)
        per_box = int(request.POST.get('per_box', 1) or 1)
        stock = stock_boxes * per_box
        min_stock_box = int(request.POST.get('min_stock_box', 0) or 0)
        min_stock = min_stock_box * per_box
        date = request.POST.get('date', timezone.now().date())

        total = sell_price * stock

        product = Product.objects.create(
            name=name,
            category_id=category_id,
            unit_id=unit_id,
            sell_price=sell_price,
            stock=stock,
            min_stock=min_stock,
            per_box=per_box,
        )

        purchase = Purchase.objects.create(
            user=request.user,
            total_amount=total,
            purchase_date=date,
            note=f'Yangi tovar: {name}',
        )
        PurchaseItem.objects.create(
            purchase=purchase,
            product=product,
            quantity=stock,
            sell_price=sell_price,
            total=total,
        )
        StockMovement.objects.create(
            product=product,
            movement_type='purchase',
            quantity=stock,
            price=sell_price,
        )

        return redirect('inventory:product_list')

    context = {
        'categories': categories,
        'units': units,
    }
    return render(request, 'inventory/add_product.html', context)


@login_required(login_url='login')
def refill_product(request):
    products = Product.objects.filter(is_active=True).select_related('unit')

    if request.method == 'POST':
        product_id = request.POST.get('product')
        quantity = int(request.POST.get('quantity', 0) or 0)
        sell_price = int(request.POST.get('sell_price', 0) or 0)
        date = request.POST.get('date', timezone.now().date())

        product = Product.objects.get(pk=product_id)

        product.stock += quantity
        if sell_price > 0:
            product.sell_price = sell_price
        product.save()

        total = product.sell_price * quantity

        purchase = Purchase.objects.create(
            user=request.user,
            total_amount=total,
            purchase_date=date,
            note=f'To\'ldirish: {product.name}',
        )
        PurchaseItem.objects.create(
            purchase=purchase,
            product=product,
            quantity=quantity,
            sell_price=product.sell_price,
            total=total,
        )
        StockMovement.objects.create(
            product=product,
            movement_type='purchase',
            quantity=quantity,
            price=product.sell_price,
        )

        return redirect('inventory:product_list')

    context = {
        'products': products,
    }
    return render(request, 'inventory/refill_product.html', context)


@login_required
def product_price_api(request, product_id):
    try:
        product = Product.objects.get(pk=product_id)
        return JsonResponse({
            'sell_price': product.sell_price,
        })
    except Product.DoesNotExist:
        return JsonResponse({'error': 'not found'}, status=404)
    

@login_required(login_url='login')
def bozor_page(request):
    from finance.models import Shop
    products = Product.objects.filter(is_active=True, stock__gt=0).select_related('category', 'unit')
    shops = Shop.objects.all()
    context = {
        'products': products,
        'shops': shops,
    }
    return render(request, 'inventory/bozor_page.html', context)

@require_POST
@login_required(login_url='login')
def bozor_send_api(request):
    try:
        data = json.loads(request.body)
        items = data.get('items', [])
        date = data.get('date')
        shop_id = data.get('shop_id')

        if not items:
            return JsonResponse({'error': 'Tovar tanlanmagan'}, status=400)
        if not shop_id:
            return JsonResponse({'error': 'Dokon tanlanmagan'}, status=400)

        from finance.models import Shop, BazarStock

        shop = Shop.objects.get(pk=shop_id)

        sale = Sale.objects.create(
            user=request.user,
            status='confirmed',
            sale_date=date,
            note=f'Bozorga jo\'natish — {shop.name}',
        )

        total_amount = 0

        for item in items:
            product = Product.objects.get(pk=item['id'])
            qty = int(item['qty'])
            price = int(item['price'])

            SaleItem.objects.create(
                sale=sale,
                product=product,
                quantity=qty,
                price=price,
                total=qty * price,
            )

            total_amount += qty * price

            Product.objects.filter(pk=product.id).update(
                stock=F('stock') - qty
            )

            StockMovement.objects.create(
                product=product,
                movement_type='sale',
                quantity=qty,
                price=price,
            )

            bazar, created = BazarStock.objects.get_or_create(
                product=product,
                shop=shop,
                defaults={'quantity': 0}
            )
            bazar.quantity += qty
            bazar.save()

        sale.total_amount = total_amount
        sale.total_cost = 0
        sale.profit = 0
        sale.save()

        return JsonResponse({'ok': True, 'sale_id': sale.id})

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)



@login_required(login_url='login')
def download_sale_excel(request, sale_id):
    sale = Sale.objects.get(pk=sale_id)
    items = sale.items.select_related('product', 'product__unit')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Bozorga ketish'

    header_font = Font(bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    bold_font = Font(bold=True, size=11)

    ws.merge_cells('A1:F1')
    ws['A1'] = f'Bozorga ketish — {sale.sale_date}'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:F2')
    ws['A2'] = f'{sale.note}'
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A2'].font = Font(size=10, color='888888')

    headers = ['№', 'Tovar', 'Korobka', 'Dona', 'Narxi (so\'m)', 'Summa (so\'m)']
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 20

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    row = 5
    for i, item in enumerate(items, 1):
        per_box = item.product.per_box if item.product.per_box > 0 else 1
        boxes = item.quantity // per_box
        box_text = f'{boxes} kor.'
        

        ws.cell(row=row, column=1, value=i).border = border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')

        ws.cell(row=row, column=2, value=item.product.name).border = border

        
        ws.cell(row=row, column=3, value=box_text).border = border
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')

        ws.cell(row=row, column=4, value=f'{item.quantity} {item.product.unit.short_name}').border = border
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')

        ws.cell(row=row, column=5, value=item.price).border = border
        ws.cell(row=row, column=5).number_format = '#,##0'
        ws.cell(row=row, column=5).alignment = Alignment(horizontal='right')

        ws.cell(row=row, column=6, value=item.total).border = border
        ws.cell(row=row, column=6).number_format = '#,##0'
        ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=6).font = bold_font

        row += 1

    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws.cell(row=row, column=1, value='JAMI:').font = Font(bold=True, size=13)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='right')
    ws.cell(row=row, column=6, value=sale.total_amount).font = Font(bold=True, size=13, color='e94560')
    ws.cell(row=row, column=6).number_format = '#,##0'
    ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')

    row += 2
    ws.cell(row=row, column=1, value=f'Sana: {sale.sale_date}').font = Font(size=10, color='888888')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'bozorga_ketish_{sale.sale_date}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@login_required(login_url='login')
def bozordagi_tovarlar(request, shop_id):
    from finance.models import BazarStock, Shop
    shop = Shop.objects.get(pk=shop_id)
    bazar_items = BazarStock.objects.filter(shop=shop, quantity__gt=0).select_related('product', 'product__category', 'product__unit')
    context = {
        'bazar_items': bazar_items,
        'shop': shop,
    }
    return render(request, 'inventory/bozordagi_tovarlar.html', context)




@require_POST
@login_required(login_url='login')
def bazar_sell_api(request):
    try:
        from finance.models import BazarStock, BazarSale, BazarSaleItem, Shop
        data = json.loads(request.body)
        items = data.get('items', [])
        date = data.get('date')
        client_name = data.get('client_name', '')
        client_phone = data.get('client_phone', '')
        payment_status = data.get('payment_status', 'paid')
        shop_id = data.get('shop_id')

        if not items:
            return JsonResponse({'error': 'Tovar tanlanmagan'}, status=400)

        shop = Shop.objects.get(pk=shop_id) if shop_id else None

        sale = BazarSale.objects.create(
            user=request.user,
            client_name=client_name,
            client_phone=client_phone,
            payment_status=payment_status,
            sale_date=date,
            shop=shop,
        )

        total_amount = 0

        for item in items:
            product = Product.objects.get(pk=item['id'])
            qty = int(item['qty'])
            price = int(item['price'])

            BazarSaleItem.objects.create(
                sale=sale,
                product=product,
                quantity=qty,
                price=price,
                total=qty * price,
            )

            total_amount += qty * price

            bazar = BazarStock.objects.get(product=product, shop=shop)
            bazar.quantity -= qty
            bazar.save()

            StockMovement.objects.create(
                product=product,
                movement_type='sale',
                quantity=qty,
                price=price,
            )

        sale.total_amount = total_amount
        sale.save()

        return JsonResponse({'ok': True, 'sale_id': sale.id})

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)
    


@login_required(login_url='login')
def qarzdorlar_page(request, shop_id):
    from finance.models import BazarSale, Shop
    from django.db.models import Sum
    shop = Shop.objects.get(pk=shop_id)
    debts = BazarSale.objects.filter(payment_status='debt', shop=shop)
    total_debt = debts.aggregate(total=Sum('total_amount'))['total'] or 0
    context = {
        'debts': debts,
        'total_debt': total_debt,
        'shop': shop,
    }
    return render(request, 'inventory/qarzdorlar.html', context)


@login_required(login_url='login')
def bazar_sale_detail_api(request, sale_id):
    from finance.models import BazarSale
    sale = BazarSale.objects.get(pk=sale_id)
    items = sale.items.select_related('product')
    data = {
        'client_name': sale.client_name,
        'date': str(sale.sale_date),
        'items': [
            {
                'product': item.product.name,
                'qty': item.quantity,
                'price': item.price,
                'total': item.total,
            }
            for item in items
        ]
    }
    return JsonResponse(data)


@require_POST
@login_required(login_url='login')
def bazar_mark_paid_api(request, sale_id):
    from finance.models import BazarSale
    try:
        sale = BazarSale.objects.get(pk=sale_id)
        sale.payment_status = 'paid'
        sale.save()
        return JsonResponse({'ok': True})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)
    

@login_required(login_url='login')
def xabar_page(request, shop_id=None):
    from finance.models import Message

    is_bozor = '/bozor/' in request.path and '/xabar/' in request.path

    shop_id = None
    if is_bozor:
        parts = request.path.split('/')
        for i, part in enumerate(parts):
            if part == 'bozor' and i + 1 < len(parts):
                try:
                    shop_id = int(parts[i + 1])
                except ValueError:
                    pass

    if request.method == 'POST':
        text = request.POST.get('text', '').strip()
        if text:
            direction = 'to_sklad' if is_bozor else 'to_bozor'
            Message.objects.create(
                sender=request.user,
                direction=direction,
                text=text,
            )
            return redirect(request.path)

    if is_bozor:
        Message.objects.filter(direction='to_bozor', is_read=False).update(is_read=True)
    else:
        Message.objects.filter(direction='to_sklad', is_read=False).update(is_read=True)

    messages_list = Message.objects.all().order_by('created_at')[:50]

    base_template = 'base_bozor.html' if is_bozor else 'base.html'

    from finance.models import Shop
    shop = Shop.objects.get(pk=shop_id) if shop_id else None

    context = {
        'messages_list': messages_list,
        'base_template': base_template,
        'shop': shop,
    }
    return render(request, 'inventory/xabar_skladga.html', context)


@login_required(login_url='login')
def xabar_count_api(request):
    from finance.models import Message
    direction = request.GET.get('direction', 'to_sklad')
    count = Message.objects.filter(direction=direction, is_read=False).count()
    return JsonResponse({'count': count})


@require_POST
@login_required(login_url='login')
def xabar_read_api(request):
    from finance.models import Message
    direction = request.POST.get('direction', 'to_sklad')
    Message.objects.filter(direction=direction, is_read=False).update(is_read=True)
    return JsonResponse({'ok': True})



@login_required(login_url='login')
def bozor_sotuvlar(request, shop_id):
    from finance.models import BazarSale, BazarSaleItem, Shop
    from django.db.models import Sum
    from datetime import date

    shop = Shop.objects.get(pk=shop_id)

    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    product_search = request.GET.get('product', '')
    status_filter = request.GET.get('status', '')

    sales = BazarSale.objects.filter(shop=shop).select_related('user')

    if date_from:
        sales = sales.filter(sale_date__gte=date_from)
    if date_to:
        sales = sales.filter(sale_date__lte=date_to)
    if status_filter:
        sales = sales.filter(payment_status=status_filter)
    if product_search:
        sale_ids = BazarSaleItem.objects.filter(
            sale__shop=shop,
            product__name__icontains=product_search
        ).values_list('sale_id', flat=True)
        sales = sales.filter(id__in=sale_ids)

    today_sales = BazarSale.objects.filter(
        shop=shop, sale_date=date.today()
    ).aggregate(total=Sum('total_amount'))['total'] or 0

    today_debts = BazarSale.objects.filter(
        shop=shop, sale_date=date.today(), payment_status='debt'
    ).aggregate(total=Sum('total_amount'))['total'] or 0

    today_paid = BazarSale.objects.filter(
        shop=shop, sale_date=date.today(), payment_status='paid'
    ).aggregate(total=Sum('total_amount'))['total'] or 0

    filtered_total = sales.aggregate(total=Sum('total_amount'))['total'] or 0

    context = {
        'sales': sales[:100],
        'shop': shop,
        'today_sales': today_sales,
        'today_debts': today_debts,
        'today_paid': today_paid,
        'filtered_total': filtered_total,
        'date_from': date_from,
        'date_to': date_to,
        'product_search': product_search,
        'status_filter': status_filter,
    }
    return render(request, 'inventory/bozor_sotuvlar.html', context)

@login_required(login_url='login')
def bozorga_ketuvlar(request):
    from django.db.models import Sum
    from datetime import date
    from dateutil.relativedelta import relativedelta

    one_year_ago = date.today() - relativedelta(years=2, months=7)
    Sale.objects.filter(note__startswith='Bozorga', sale_date__lt=one_year_ago).delete()

    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    month = request.GET.get('month', '')

    sales = Sale.objects.filter(note__startswith='Bozorga').select_related('user')

    month_names = {
        1: 'Yanvar', 2: 'Fevral', 3: 'Mart', 4: 'Aprel',
        5: 'May', 6: 'Iyun', 7: 'Iyul', 8: 'Avgust',
        9: 'Sentabr', 10: 'Oktabr', 11: 'Noyabr', 12: 'Dekabr'
    }

    if month:
        year, m = month.split('-')
        year = int(year)
        m = int(m)
        sales = sales.filter(sale_date__year=year, sale_date__month=m)
        selected_label = f'{month_names[m]} {year}'
        selected_sum = sales.aggregate(total=Sum('total_amount'))['total'] or 0
    elif date_from or date_to:
        if date_from:
            sales = sales.filter(sale_date__gte=date_from)
        if date_to:
            sales = sales.filter(sale_date__lte=date_to)
        selected_label = f'{date_from or "..."} — {date_to or "..."}'
        selected_sum = sales.aggregate(total=Sum('total_amount'))['total'] or 0
    else:
        selected_label = 'Bugun'
        selected_sum = Sale.objects.filter(
            note__startswith='Bozorga',
            sale_date=date.today()
        ).aggregate(total=Sum('total_amount'))['total'] or 0

    this_month_sum = Sale.objects.filter(
        note__startswith='Bozorga',
        sale_date__year=date.today().year,
        sale_date__month=date.today().month,
    ).aggregate(total=Sum('total_amount'))['total'] or 0

    months_list = []
    current = date.today()
    for i in range(7):
        m = current.month + i
        y = current.year
        if m > 12:
            m -= 12
            y += 1
        key = f'{y}-{m:02d}'
        label = f'{month_names[m]} {y}'
        months_list.append({'key': key, 'label': label})

    context = {
        'sales': sales[:100],
        'selected_label': selected_label,
        'selected_sum': selected_sum,
        'this_month_sum': this_month_sum,
        'date_from': date_from,
        'date_to': date_to,
        'current_month': month,
        'months_list': months_list,
    }
    return render(request, 'inventory/bozorga_ketuvlar.html', context)


@require_POST
@login_required(login_url='login')
def delete_old_records(request):
    password = request.POST.get('password', '')
    if password != 'admin777':
        return JsonResponse({'error': 'Parol noto\'g\'ri!'}, status=403)

    year = int(request.POST.get('year', 0))
    month_num = int(request.POST.get('month', 0))

    if not year or not month_num:
        return JsonResponse({'error': 'Oy va yilni tanlang!'}, status=400)

    count, _ = Sale.objects.filter(
        note__startswith='Bozorga',
        sale_date__year=year,
        sale_date__month=month_num,
    ).delete()

    return JsonResponse({'ok': True, 'deleted': count})


@login_required(login_url='login')
def dashboard(request):
    # Проверка пароля
    if not request.session.get('dashboard_verified'):
        if request.method == 'POST' and 'dash_password' in request.POST:
            if request.POST.get('dash_password') == 'admin777':
                request.session['dashboard_verified'] = True
            else:
                return render(request, 'inventory/dashboard_password.html', {
                    'error': 'Parol noto\'g\'ri!'
                })
        else:
            return render(request, 'inventory/dashboard_password.html', {})
    
   

    today = date.today()
    week_ago = today - timedelta(days=7)

    all_products = Product.objects.filter(is_active=True)
    sklad_value = sum(p.stock * p.sell_price for p in all_products)
    sklad_count = all_products.count()
    sklad_items = sum(p.stock for p in all_products)

    kirim_today = Purchase.objects.filter(
        purchase_date=str(today),
    ).aggregate(total=Sum('total_amount'))['total'] or 0

    kirim_month = Purchase.objects.filter(
        purchase_date__year=today.year,
        purchase_date__month=today.month,
    ).aggregate(total=Sum('total_amount'))['total'] or 0

    shops = Shop.objects.all()
    shop_ids = list(shops.values_list('id', flat=True))

    bozor_today = BazarSale.objects.filter(
        sale_date=today, shop_id__in=shop_ids
    ).aggregate(total=Sum('total_amount'))['total'] or 0

    bozor_week = BazarSale.objects.filter(
        sale_date__gte=week_ago, shop_id__in=shop_ids
    ).aggregate(total=Sum('total_amount'))['total'] or 0

    bozor_month = BazarSale.objects.filter(
        sale_date__year=today.year,
        sale_date__month=today.month,
        shop_id__in=shop_ids,
    ).aggregate(total=Sum('total_amount'))['total'] or 0

    total_debts = BazarSale.objects.filter(
        payment_status='debt', shop_id__in=shop_ids
    ).aggregate(total=Sum('total_amount'))['total'] or 0

    ketuvlar_month = Sale.objects.filter(
        note__startswith='Bozorga',
        sale_date__year=today.year,
        sale_date__month=today.month,
    ).aggregate(total=Sum('total_amount'))['total'] or 0

    shop_stats = []
    for shop in shops:
        shop_today = BazarSale.objects.filter(
            shop=shop, sale_date=today
        ).aggregate(total=Sum('total_amount'))['total'] or 0

        shop_week = BazarSale.objects.filter(
            shop=shop, sale_date__gte=week_ago
        ).aggregate(total=Sum('total_amount'))['total'] or 0

        shop_month = BazarSale.objects.filter(
            shop=shop, sale_date__year=today.year, sale_date__month=today.month,
        ).aggregate(total=Sum('total_amount'))['total'] or 0

        shop_debts = BazarSale.objects.filter(
            shop=shop, payment_status='debt'
        ).aggregate(total=Sum('total_amount'))['total'] or 0

        shop_stats.append({
            'id': shop.id,
            'name': shop.name,
            'today': shop_today,
            'week': shop_week,
            'month': shop_month,
            'debts': shop_debts,
        })

    chart_labels_30 = []
    chart_data_30 = []
    shop_chart_data_30 = {}
    for shop in shops:
        shop_chart_data_30[shop.name] = []

    for i in range(29, -1, -1):
        d = today - timedelta(days=i)
        chart_labels_30.append(d.strftime('%d.%m'))
        day_sum = BazarSale.objects.filter(
            sale_date=d, shop_id__in=shop_ids
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        chart_data_30.append(day_sum)

        for shop in shops:
            s_sum = BazarSale.objects.filter(
                shop=shop, sale_date=d
            ).aggregate(total=Sum('total_amount'))['total'] or 0
            shop_chart_data_30[shop.name].append(s_sum)

    chart_labels_7 = []
    chart_data_7 = []
    shop_chart_data_7 = {}
    for shop in shops:
        shop_chart_data_7[shop.name] = []

    for i in range(6, -1, -1):
        d = today - timedelta(days=i)
        chart_labels_7.append(d.strftime('%d.%m'))
        day_sum = BazarSale.objects.filter(
            sale_date=d, shop_id__in=shop_ids
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        chart_data_7.append(day_sum)

        for shop in shops:
            s_sum = BazarSale.objects.filter(
                shop=shop, sale_date=d
            ).aggregate(total=Sum('total_amount'))['total'] or 0
            shop_chart_data_7[shop.name].append(s_sum)

    context = {
        'sklad_value': sklad_value,
        'sklad_count': sklad_count,
        'sklad_items': sklad_items,
        'kirim_month': kirim_month,
        'kirim_today': kirim_today,
        'bozor_today': bozor_today,
        'bozor_week': bozor_week,
        'bozor_month': bozor_month,
        'total_debts': total_debts,
        'ketuvlar_month': ketuvlar_month,
        'shop_stats': shop_stats,
        'chart_labels_30': chart_labels_30,
        'chart_data_30': chart_data_30,
        'shop_chart_data_30': shop_chart_data_30,
        'chart_labels_7': chart_labels_7,
        'chart_data_7': chart_data_7,
        'shop_chart_data_7': shop_chart_data_7,
    }
    return render(request, 'inventory/dashboard.html', context)



@login_required(login_url='login')
def product_stats_api(request, product_id):
    from finance.models import BazarSaleItem, Shop
    from django.db.models import Sum
    from datetime import date

    today = date.today()

    month_names = {
        1: 'Yanvar', 2: 'Fevral', 3: 'Mart', 4: 'Aprel',
        5: 'May', 6: 'Iyun', 7: 'Iyul', 8: 'Avgust',
        9: 'Sentabr', 10: 'Oktabr', 11: 'Noyabr', 12: 'Dekabr'
    }

    product = Product.objects.get(pk=product_id)
    per_box = product.per_box if product.per_box > 0 else 1
    shops = Shop.objects.all()

    month_sales = BazarSaleItem.objects.filter(
        product=product,
        sale__sale_date__year=today.year,
        sale__sale_date__month=today.month,
    )
    month_qty = month_sales.aggregate(total=Sum('quantity'))['total'] or 0
    month_sum = month_sales.aggregate(total=Sum('total'))['total'] or 0
    month_boxes = month_qty // per_box

    shops_month = []
    for shop in shops:
        s_qty = BazarSaleItem.objects.filter(
            product=product,
            sale__shop=shop,
            sale__sale_date__year=today.year,
            sale__sale_date__month=today.month,
        ).aggregate(total=Sum('quantity'))['total'] or 0
        s_sum = BazarSaleItem.objects.filter(
            product=product,
            sale__shop=shop,
            sale__sale_date__year=today.year,
            sale__sale_date__month=today.month,
        ).aggregate(total=Sum('total'))['total'] or 0
        shops_month.append({
            'name': shop.name,
            'qty': s_qty,
            'boxes': s_qty // per_box,
            'total': s_sum,
        })

    months_data = []
    start_year = 2026
    start_month = 3
    total_months = (today.year - start_year) * 12 + (today.month - start_month) + 1 + 6

    for i in range(total_months):
        m = start_month + i
        y = start_year
        while m > 12:
            m -= 12
            y += 1

        m_qty = BazarSaleItem.objects.filter(
            product=product,
            sale__sale_date__year=y,
            sale__sale_date__month=m,
        ).aggregate(total=Sum('quantity'))['total'] or 0
        m_sum = BazarSaleItem.objects.filter(
            product=product,
            sale__sale_date__year=y,
            sale__sale_date__month=m,
        ).aggregate(total=Sum('total'))['total'] or 0

        m_shops = []
        for shop in shops:
            s_qty = BazarSaleItem.objects.filter(
                product=product,
                sale__shop=shop,
                sale__sale_date__year=y,
                sale__sale_date__month=m,
            ).aggregate(total=Sum('quantity'))['total'] or 0
            m_shops.append({
                'name': shop.name,
                'qty': s_qty,
                'boxes': s_qty // per_box,
            })

        months_data.append({
            'month': month_names[m],
            'year': y,
            'qty': m_qty,
            'boxes': m_qty // per_box,
            'total': m_sum,
            'shops': m_shops,
            'is_current': (y == today.year and m == today.month),
        })

    months_data.reverse()

    data = {
        'name': product.name,
        'per_box': per_box,
        'stock': product.stock,
        'stock_boxes': product.stock // per_box,
        'current_month': month_names[today.month],
        'month_qty': month_qty,
        'month_boxes': month_boxes,
        'month_sum': month_sum,
        'shops_month': shops_month,
        'months': months_data,
    }
    return JsonResponse(data)





@login_required(login_url='login')
def monthly_sales(request):
    if not request.session.get('dashboard_verified'):
        return redirect('inventory:dashboard')
    

    today = date.today()
    shops = Shop.objects.all()

    month_names = {
        1: 'Yanvar', 2: 'Fevral', 3: 'Mart', 4: 'Aprel',
        5: 'May', 6: 'Iyun', 7: 'Iyul', 8: 'Avgust',
        9: 'Sentabr', 10: 'Oktabr', 11: 'Noyabr', 12: 'Dekabr'
    }

    start_year = 2026
    start_month = 3
    total_months = (today.year - start_year) * 12 + (today.month - start_month) + 1 + 6

    months_data = []
    for i in range(total_months):
        m = start_month + i
        y = start_year
        while m > 12:
            m -= 12
            y += 1

        total = BazarSale.objects.filter(
            sale_date__year=y, sale_date__month=m,
        ).aggregate(total=Sum('total_amount'))['total'] or 0

        shop_totals = []
        for shop in shops:
            s_total = BazarSale.objects.filter(
                shop=shop, sale_date__year=y, sale_date__month=m,
            ).aggregate(total=Sum('total_amount'))['total'] or 0
            shop_totals.append({
                'name': shop.name,
                'total': s_total,
            })

        months_data.append({
            'month': month_names[m],
            'year': y,
            'total': total,
            'shops': shop_totals,
            'is_current': (y == today.year and m == today.month),
        })

    months_data.reverse()

    context = {
        'months_data': months_data,
        'shops': shops,
    }
    return render(request, 'inventory/monthly_sales.html', context)



@login_required(login_url='login')
def bazar_add_product(request, shop_id):
    from finance.models import BazarStock, Shop
    shop = Shop.objects.get(pk=shop_id)
    categories = Category.objects.all()
    units = Unit.objects.all()

    if request.method == 'POST':
        name = request.POST.get('name')
        category_id = request.POST.get('category')
        unit_id = request.POST.get('unit')
        sell_price = int(request.POST.get('sell_price', 0) or 0)
        stock_boxes = int(request.POST.get('stock', 0) or 0)
        per_box = int(request.POST.get('per_box', 1) or 1)
        stock = stock_boxes * per_box

        product = Product.objects.create(
            name=name,
            category_id=category_id,
            unit_id=unit_id,
            sell_price=sell_price,
            stock=0,
            min_stock=0,
            per_box=per_box,
        )

        BazarStock.objects.create(
            product=product,
            shop=shop,
            quantity=stock,
        )

        return redirect(f'/bozor/{shop_id}/')

    context = {
        'categories': categories,
        'units': units,
        'shop': shop,
    }
    return render(request, 'inventory/bazar_add_product.html', context)